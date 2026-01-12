# ======================================================================
# IMPORTS — WHAT THEY ARE AND WHY THEY ARE USED
# ======================================================================

import random
# We use random.uniform() to pick a random delay between requests so our
# traffic pattern is less "bot-like" and more polite.

import time
# time.sleep() actually performs the wait between requests.

from datetime import datetime
# datetime.now() is used to generate timestamps for file names, so you
# can see exactly when each output was created.

from zoneinfo import ZoneInfo
# This lets us pin timestamps to a specific time zone ("America/Chicago")
# instead of using the machine's local time, which could change.

import pandas as pd
# Pandas is used to:
#   - Store all player rows in a DataFrame
#   - Drop duplicates
#   - Sort data
#   - Export the results to an Excel file.

from bs4 import BeautifulSoup
# BeautifulSoup turns HTML into a tree of tags that we can navigate.
# We use it to find the players table and pull out the data we care about.

from playwright.sync_api import sync_playwright
# Playwright controls a real browser (Chromium). This is important
# because Basketball-Reference often hides tables or uses dynamic content,
# and a simple requests.get() might not see the full DOM.

from openpyxl.utils import get_column_letter
# Used when "autofitting" Excel columns; we need to map 1 -> "A", 2 -> "B", etc.

from openpyxl.styles import Alignment
# Lets us control cell alignment. We use it to left-align the header row.

from openpyxl.worksheet.views import Selection
# Used to set which cell is "active" when you open the workbook (A1).


# ======================================================================
# USER CONFIGURATION — YOU CAN EDIT THESE VALUES
# ======================================================================

BASE_URL = "https://www.basketball-reference.com"

# Starting and ending letters to scrape.
# The script will automatically fix reversed ranges (e.g., START=B, END=A).
START_LETTER = "a"
END_LETTER = "z"

# Decide which outputs you want the script to produce.
INCLUDE_HTML = True        # Save all raw HTML for A–Z into a text file.
INCLUDE_PLAYER_DATA = True  # Parse player table into Excel.
INCLUDE_SQL_SCRIPT = True  # Generate a SQLite-friendly SQL script.

# Delay range between requests (seconds).
MIN_DELAY_SECONDS = 1.0
MAX_DELAY_SECONDS = 2.5

# Name of the table in the generated SQL script.
SQL_TABLE_NAME = "players"


# ======================================================================
# LETTER RANGE HANDLING
# ======================================================================

def clamp_letters(start_letter_input: str, end_letter_input: str) -> tuple[str, str]:
    """
    Clean up and validate the start/end letters.

    - Ensure they are single characters.
    - Force them into 'a'..'z'.
    - If the order is reversed (e.g., 'c' to 'a'), swap them to 'a'..'c'.

    This function prevents bad input from breaking the rest of the script.
    """
    # Normalize input: default to "a" / "z", remove spaces, lower-case, take first char.
    s = (start_letter_input or "a").strip().lower()[:1]
    e = (end_letter_input or "z").strip().lower()[:1]

    # If out of range, clamp to boundaries.
    if not ("a" <= s <= "z"):
        s = "a"
    if not ("a" <= e <= "z"):
        e = "z"

    # If the end is alphabetically before the start, swap them.
    if e < s:
        s, e = e, s

    return s, e


def build_letter_range_list(start_letter: str, end_letter: str) -> list[str]:
    """
    Build a list of letters from start to end inclusive.

    Example:
        start_letter = 'a', end_letter = 'c'
        -> ['a', 'b', 'c']
    """
    return [chr(code) for code in range(ord(start_letter), ord(end_letter) + 1)]


# ======================================================================
# POLITE DELAY
# ======================================================================

def wait_politely():
    """
    Pause for a random amount of time between MIN_DELAY_SECONDS and
    MAX_DELAY_SECONDS.

    This helps:
      - Avoid sending too many requests too quickly.
      - Look less like an abusive bot.
    """
    time.sleep(random.uniform(MIN_DELAY_SECONDS, MAX_DELAY_SECONDS))


# ======================================================================
# URL BUILDER
# ======================================================================

def build_players_letter_url(letter: str) -> str:
    """
    Given a letter like 'a', return the Basketball-Reference URL for
    the players whose last names start with that letter.

    Example:
        'a' -> 'https://www.basketball-reference.com/players/a/'
    """
    return f"{BASE_URL}/players/{letter}/"


# ======================================================================
# SAFE CONVERSION UTILITIES
# ======================================================================

def safe_to_float(value: str):
    """
    Attempt to convert a string to float.
    Return None if conversion fails.

    This prevents exceptions when parsing slightly messy or missing data.
    """
    try:
        return float(value)
    except Exception:
        return None


def safe_to_int(value: str):
    """
    Attempt to convert a string to int.
    Return None if conversion fails.
    """
    try:
        return int(value)
    except Exception:
        return None


def inches_to_meters(inches):
    """
    Convert inches to meters, rounded to 3 decimal places.

    If inches is None, return None (allows consistent 'no data' handling).
    """
    return round(inches * 0.0254, 3) if inches is not None else None


def pounds_to_kilograms(lb):
    """
    Convert pounds to kilograms, rounded to 3 decimal places.

    If lb is None, return None.
    """
    return round(lb * 0.45359237, 3) if lb is not None else None


def normalize_height_feet_inches(h):
    """
    Normalize height strings like "6-9" into "6-09" so they are consistent.

    This helps:
      - Data look cleaner.
      - Sorting behave more predictably ("6-09" vs "6-10", etc.).
    """
    if not h or "-" not in h:
        return h
    ft, inch = h.split("-", 1)
    return f"{ft}-{inch.zfill(2)}"


# ======================================================================
# HTML PARSER FOR THE PLAYERS TABLE
# ======================================================================

def parse_players_table_html(table_html: str, letter: str) -> list[dict]:
    """
    Given the HTML for the players table for a single letter, extract
    each row and return a list of dictionaries (one per player).

    This is where we translate "raw HTML" into structured data.
    """

    soup = BeautifulSoup(table_html, "html.parser")
    tbody = soup.find("tbody")
    if not tbody:
        return []

    rows = []

    # Each row in <tbody> corresponds to a player.
    for tr in tbody.find_all("tr", recursive=False):

        # The "player" cell contains:
        #   - the main link for the player
        #   - the Hall of Fame star
        #   - boldness for "Active"
        player_cell = tr.find(["th", "td"], attrs={"data-stat": "player"})
        if not player_cell:
            continue

        # Unique Basketball-Reference ID (e.g., "jamesle01").
        unique_id = player_cell.get("data-append-csv", "") or ""

        # Anchor tag with the link to the player's page.
        link_tag = player_cell.find("a")
        if not link_tag or not link_tag.get("href"):
            continue

        player_name = link_tag.get_text(strip=True)
        player_url = BASE_URL + link_tag["href"]

        # "*" in the text indicates Hall of Fame.
        cell_text = player_cell.get_text(strip=True)
        hall_of_fame = "x" if "*" in cell_text else ""

        # Active players are bolded via <strong>.
        status = "Active" if player_cell.find("strong") else "Inactive"

        # Helper functions for cells.
        def get_text(stat):
            td = tr.find("td", attrs={"data-stat": stat})
            return td.get_text(strip=True) if td else ""

        def get_attr(stat, attr):
            td = tr.find("td", attrs={"data-stat": stat})
            return td.get(attr) if td else None

        # Basic career info.
        year_start = get_text("year_min")
        year_end = get_text("year_max")
        pos = get_text("pos")

        # Height (string and normalized), plus inches and meters.
        height_ft_in_raw = get_text("height")
        height_ft_in = normalize_height_feet_inches(height_ft_in_raw)

        height_in = safe_to_float(get_attr("height", "csk") or "")
        height_m = inches_to_meters(height_in)

        # Weight in lb and kg.
        weight_lb = safe_to_int(get_text("weight"))
        weight_kg = pounds_to_kilograms(weight_lb)

        # Birth date: short (YYYY.MM.DD) and long formats.
        birth_csk = (get_attr("birth_date", "csk") or "").strip()
        birthday_short = (
            f"{birth_csk[0:4]}.{birth_csk[4:6]}.{birth_csk[6:8]}"
            if len(birth_csk) == 8
            else ""
        )
        birthday_long = get_text("birth_date")

        # Colleges: may contain multiple <a> links; we join with " | ".
        colleges_td = tr.find("td", attrs={"data-stat": "colleges"})
        colleges = ""
        if colleges_td:
            colleges = " | ".join(
                a.get_text(strip=True) for a in colleges_td.find_all("a")
            )

        rows.append({
            "letter": letter.upper(),
            "unique_id": unique_id,
            "player_name": player_name,
            "player_url": player_url,
            "HOF": hall_of_fame,
            "status": status,
            "year_start": year_start,
            "year_end": year_end,
            "pos": pos,
            "height_in": height_in,
            "height_ft_in": height_ft_in,
            "height_m": height_m,
            "weight_lb": weight_lb,
            "weight_kg": weight_kg,
            "birthday": birthday_short,
            "birthday_long": birthday_long,
            "colleges": colleges,
        })

    return rows


# ======================================================================
# HTML FILE HEADER WRITER
# ======================================================================

def write_html_file_header(fh, title, start_letter, end_letter, timestamp):
    """
    Write a small text header into the combined HTML dump file so you
    know what letters it covers and when it was generated.
    """
    fh.write(f"{title}\n")
    fh.write(f"Letters: {start_letter.upper()}-{end_letter.upper()}\n")
    fh.write(f"Generated: {timestamp} America/Chicago\n\n")


# ======================================================================
# SQL ESCAPING
# ======================================================================

def sql_escape(value):
    """
    Escape a Python value for safe use in a SQLite INSERT statement.

    - If value is None, return the literal SQL NULL (no quotes).
    - Otherwise, convert to string, escape any apostrophes by doubling them,
      and wrap in single quotes.

    Example:
      "O'Brien" -> 'O''Brien'
    """
    if value is None:
        return "NULL"
    text = str(value).replace("'", "''")
    return f"'{text}'"


# ======================================================================
# SQL SCRIPT GENERATOR
# ======================================================================

def generate_sql_script(rows: list[dict], table_name: str) -> str:
    """
    Build a full SQLite SQL script that:

      1. Drops the table if it exists.
      2. Creates the table with columns matching our data.
      3. Inserts every row of scraped player data.

    The result is a single string you can write to a .txt file and later
    execute in SQLite.
    """

    sql_lines = []

    # Drop any old version of the table so the script is idempotent.
    sql_lines.append(f"DROP TABLE IF EXISTS {table_name};")
    sql_lines.append("")

    # Create the table with columns aligned to our output schema.
    sql_lines.append(f"CREATE TABLE {table_name} (")
    sql_lines.append("    letter TEXT,")
    sql_lines.append("    unique_id TEXT,")
    sql_lines.append("    player_name TEXT,")
    sql_lines.append("    player_url TEXT,")
    sql_lines.append("    HOF TEXT,")
    sql_lines.append("    status TEXT,")
    sql_lines.append("    year_start TEXT,")
    sql_lines.append("    year_end TEXT,")
    sql_lines.append("    pos TEXT,")
    sql_lines.append("    height_in REAL,")
    sql_lines.append("    height_ft_in TEXT,")
    sql_lines.append("    height_m REAL,")
    sql_lines.append("    weight_lb INTEGER,")
    sql_lines.append("    weight_kg REAL,")
    sql_lines.append("    birthday TEXT,")
    sql_lines.append("    birthday_long TEXT,")
    sql_lines.append("    colleges TEXT")
    sql_lines.append(");")
    sql_lines.append("")

    # Generate one INSERT statement per row.
    for row in rows:
        insert_line = (
            f"INSERT INTO {table_name} VALUES ("
            f"{sql_escape(row['letter'])}, "
            f"{sql_escape(row['unique_id'])}, "
            f"{sql_escape(row['player_name'])}, "
            f"{sql_escape(row['player_url'])}, "
            f"{sql_escape(row['HOF'])}, "
            f"{sql_escape(row['status'])}, "
            f"{sql_escape(row['year_start'])}, "
            f"{sql_escape(row['year_end'])}, "
            f"{sql_escape(row['pos'])}, "
            f"{row['height_in'] if row['height_in'] is not None else 'NULL'}, "
            f"{sql_escape(row['height_ft_in'])}, "
            f"{row['height_m'] if row['height_m'] is not None else 'NULL'}, "
            f"{row['weight_lb'] if row['weight_lb'] is not None else 'NULL'}, "
            f"{row['weight_kg'] if row['weight_kg'] is not None else 'NULL'}, "
            f"{sql_escape(row['birthday'])}, "
            f"{sql_escape(row['birthday_long'])}, "
            f"{sql_escape(row['colleges'])}"
            f");"
        )
        sql_lines.append(insert_line)

    return "\n".join(sql_lines)


# ======================================================================
# MAIN SCRIPT
# ======================================================================

def main():
    """
    Orchestrate the full process:

      1. Normalize and validate letter range.
      2. Build output file names.
      3. Use Playwright to visit each letter page.
      4. Optionally dump combined HTML.
      5. Optionally parse player data into Excel.
      6. Optionally generate a SQLite SQL script.
    """

    # Step 1: Normalize letters (fix reversed inputs).
    start_letter, end_letter = clamp_letters(START_LETTER, END_LETTER)

    # Step 2: Build a timestamp string for filenames.
    timestamp = datetime.now(ZoneInfo("America/Chicago")).strftime(
        "%Y.%m.%d-%H.%M.%S"
    )

    # Step 3: Construct base filename so all outputs line up neatly in a folder.
    base_filename = (
        f"BR Players A-Z ({start_letter.upper()}-{end_letter.upper()}) "
        f"as of {timestamp}"
    )

    # Derived filenames for each kind of output.
    html_filename = f"{base_filename} HTML.txt"
    excel_filename = f"{base_filename}.xlsx"
    sql_filename = f"{base_filename} SQL SCRIPT.txt"

    # Build the list of letters to process.
    letters = build_letter_range_list(start_letter, end_letter)

    # This list will store all parsed player rows from all letters.
    all_rows = []

    # Step 4: Use Playwright to browse the site.
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page()

        # NOTE: We intentionally do NOT use page.route() here.
        #   - Using route() caused noisy async CancelledError logs when the
        #     browser shut down.
        #   - We remove it to keep the console clean and error-free, at the
        #     cost of downloading images/CSS/fonts (minimal impact).

        # Open the HTML dump file if requested.
        html_file = open(html_filename, "w",
                         encoding="utf-8") if INCLUDE_HTML else None

        # If we have an HTML file, write a header at the top.
        if html_file:
            write_html_file_header(
                html_file,
                "Basketball-Reference players A-Z pages",
                start_letter,
                end_letter,
                timestamp,
            )

        try:
            total = len(letters)

            for i, letter in enumerate(letters, start=1):

                url = build_players_letter_url(letter)
                print(f"[{i}/{total}] Downloading {letter.upper()}: {url}")

                # Load the page for this letter.
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(500)  # Small extra wait for stability.

                # --------------------------
                # OPTIONAL: SAVE RAW HTML
                # --------------------------
                if html_file:
                    html = page.content()
                    html_file.write("\n" + "=" * 120 + "\n")
                    html_file.write(
                        f"BEGIN LETTER {letter.upper()} | URL {url}\n")
                    html_file.write("=" * 120 + "\n")
                    html_file.write(html)
                    html_file.write("\n" + "-" * 120 + "\n")
                    html_file.write(f"END LETTER {letter.upper()}\n")
                    html_file.write("-" * 120 + "\n")

                # --------------------------
                # OPTIONAL: PARSE PLAYER DATA
                # --------------------------
                if INCLUDE_PLAYER_DATA:
                    # Wait until the players table is attached to the DOM.
                    page.wait_for_selector(
                        "table#players",
                        state="attached",
                        timeout=60000,
                    )

                    # Pull just the <table> HTML and feed it to BeautifulSoup.
                    table_html = page.locator("table#players").evaluate(
                        "el => el.outerHTML"
                    )

                    rows = parse_players_table_html(table_html, letter)
                    print(f"{letter.upper()}: extracted {len(rows)} players")

                    all_rows.extend(rows)

                # Respectful delay before the next letter.
                wait_politely()

        finally:
            # Ensure resources are closed even if something goes wrong.
            if html_file:
                html_file.close()
            browser.close()

    # ==================================================================
    # EXCEL OUTPUT
    # ==================================================================
    if INCLUDE_PLAYER_DATA and all_rows:

        # Create a DataFrame from all collected rows.
        df = pd.DataFrame(all_rows)

        # Drop duplicate players by unique_id and sort by letter + name.
        df = (
            df.drop_duplicates(subset=["unique_id"])
              .sort_values(["letter", "player_name"])
              .reset_index(drop=True)
        )

        # Ensure columns are in the desired order (and only use those that exist).
        columns = [
            "letter",
            "unique_id",
            "player_name",
            "player_url",
            "HOF",
            "status",
            "year_start",
            "year_end",
            "pos",
            "height_in",
            "height_ft_in",
            "height_m",
            "weight_lb",
            "weight_kg",
            "birthday",
            "birthday_long",
            "colleges",
        ]
        df = df[[c for c in columns if c in df.columns]]

        # Write to Excel using openpyxl so we can format.
        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:

            df.to_excel(writer, index=False, sheet_name="NBA Players")

            ws = writer.sheets["NBA Players"]

            # Add filter to header row.
            ws.auto_filter.ref = ws.dimensions

            # Left-align header text.
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="left")

            # Autofit all columns by scanning their max text length.
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(
                    col_idx)].width = max_len + 2

            # Make A1 the active cell when the workbook is opened.
            ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

        print(f"Saved Excel file: {excel_filename}")

    # ==================================================================
    # SQL OUTPUT
    # ==================================================================
    if INCLUDE_SQL_SCRIPT and all_rows:

        # Build the SQL script text.
        sql_script_text = generate_sql_script(all_rows, SQL_TABLE_NAME)

        # Write it to a .txt file (not .sql, per your preference).
        with open(sql_filename, "w", encoding="utf-8") as f:
            f.write(sql_script_text)

        print(f"Saved SQL script: {sql_filename}")

    # ==================================================================
    # HTML OUTPUT MESSAGE
    # ==================================================================
    if INCLUDE_HTML:
        print(f"Saved combined HTML file: {html_filename}")


# ======================================================================
# ENTRY POINT
# ======================================================================

if __name__ == "__main__":
    main()
